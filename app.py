# -*- coding: utf-8 -*-
"""
交通費・出張見積もりアプリ（Streamlit Cloud公開版）
テンプレートExcelに値を書き込む方式。

[修正履歴]
- get_navitime_fastest_route: 運賃計算で unit_0 + unit_3 を合算していたバグを修正。
  NAVITIMEのfareオブジェクトは支払い方法・料金体系ごとに複数の組が並列で入っており、
  単純に足し合わせてよいものではない。NAVITIMEが提供する reference_fare
  (lowest_total_ticket / lowest_total_ic) を最優先で採用し、無ければ
  unit_0（運賃）+ unit_1（特急料金）の正しい組み合わせにフォールバックする。
  （旧: unit_0 + unit_3 → 無関係な料金体系同士を合算し、実際より高い金額になっていた）
- build_best_route_patterns: excel_data に詰める運賃が「片道」のままだったバグを修正。
  Excelテンプレートの行9〜11は「電車・新幹線（往復）」「飛行機（往復）」等、
  いずれも往復を前提としたラベルであり、Q列の合計式（=M×O×P）は
  「M(=N の切り上げ)に往復1人分の運賃が入っている」前提で人数(P)を掛けているだけ。
  往復を意味する係数がテンプレート側のどこにも存在しないため、
  Python側で片道運賃を明示的に2倍してから書き込む必要がある。
"""

import json
import math
import io
import re
import time
import requests
from datetime import datetime, timedelta
import streamlit as st

from google import genai
from google.genai import types

import openpyxl
from openpyxl.utils import get_column_letter

# ============================================================
# 定数
# ============================================================

STATION_COORDS = {
    "大宮駅": (35.906702, 139.623596),
    "淀屋橋駅": (34.693729, 135.499814),
}

HOTEL_COST_PER_NIGHT_PER_PERSON = 20_000
RENTAL_CAR_COST_PER_DAY = 12_000
TAXI_FARE_PER_KM = 400
TAXI_WALK_THRESHOLD_MIN = 15

NAVITIME_URL = "https://navitime-route-totalnavi.p.rapidapi.com/route_transit"
NAVITIME_HOST = "navitime-route-totalnavi.p.rapidapi.com"
GSI_GEOCODE_URL = "https://msearch.gsi.go.jp/address-search/AddressSearch"

FLIGHT_MOVE_TYPES = {
    "plane", "flight", "air", "airplane", "aeroplane",
    "domestic_flight", "international_flight", "domestic_air"
}

GEMINI_MODEL = "gemini-3.5-flash"
MAX_RETRIES = 3
RETRY_WAIT_SECONDS = [5, 15, 30]
RETRYABLE_STATUS_CODES = {429, 503}
DEBUG_MODE = True

# テンプレートファイル
# 同じ階層のxlsxはファイル名を問わず自動探索するため、固定ファイル名の指定は不要
TEMPLATE_SHEET_NAME = "交通費"


@st.cache_data(ttl=600, show_spinner=False)
def fetch_template_from_github(raw_url: str, token: str = ""):
    """
    GitHub上のテンプレートExcelを取得する。
    raw_url は raw.githubusercontent.com 形式のURL
      例: https://raw.githubusercontent.com/<owner>/<repo>/<branch>/<path>/xxx.xlsx
    プライベートリポジトリの場合は token（GitHub Personal Access Token, repo権限）を渡すと
    Authorization ヘッダ付きで取得する。
    st.cache_data により、同一URLへの再取得は10分間キャッシュされる。
    """
    headers = {}
    if token:
        headers["Authorization"] = f"token {token}"
    res = requests.get(raw_url, headers=headers, timeout=15)
    if res.status_code != 200:
        raise ValueError(
            f"GitHubからのテンプレート取得に失敗しました（HTTP {res.status_code}）。"
            f"URLが正しいか、プライベートリポジトリならトークンが正しいか確認してください。\nURL: {raw_url}"
        )
    return res.content


def resolve_template_bytes():
    """
    テンプレートExcelのバイト列を取得する。優先順位:
    1. サイドバーで手動アップロードされたもの（session_state）
    2. secrets.toml の [template] github_raw_url で指定されたGitHub上のファイル
    3. 同じ階層にある既知のファイル名候補（ローカル運用向けのフォールバック）
    4. どれもなければ None（呼び出し側でエラー表示）

    戻り値: (bytes|None, source: str, log: list[str])
    """
    log = []

    uploaded = st.session_state.get("template_bytes")
    if uploaded:
        log.append("サイドバーで手動アップロードされたファイルを使用")
        return uploaded, "manual_upload", log

    try:
        template_secrets = st.secrets.get("template", {})
    except Exception as e:
        template_secrets = {}
        log.append(f"st.secrets読み込み自体に失敗: {e}")

    github_url = template_secrets.get("github_raw_url", "") if template_secrets else ""
    github_token = template_secrets.get("github_token", "") if template_secrets else ""

    if not template_secrets:
        log.append("secrets.tomlに [template] セクションが見つかりません（未設定）")
    elif not github_url:
        log.append("secrets.tomlの [template] に github_raw_url が見つかりません（キー名の誤字の可能性）")

    if github_url:
        log.append(f"GitHub raw URLを検出: {github_url}")
        try:
            content = fetch_template_from_github(github_url, github_token)
            log.append("GitHubからの取得に成功")
            return content, "github", log
        except Exception as e:
            log.append(f"GitHubからの取得に失敗: {e}")

    import os, glob
    base_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_paths = sorted(glob.glob(os.path.join(base_dir, "*.xlsx")))
    log.append(f"{base_dir} 内で見つかった .xlsx ファイル: "
               f"{[os.path.basename(p) for p in xlsx_paths] if xlsx_paths else '(なし)'}")

    for path in xlsx_paths:
        try:
            wb_check = openpyxl.load_workbook(path, read_only=True)
            sheetnames = wb_check.sheetnames
            wb_check.close()
        except Exception as e:
            log.append(f"「{os.path.basename(path)}」の読み込みに失敗: {e}")
            continue
        if TEMPLATE_SHEET_NAME in sheetnames:
            log.append(f"「{TEMPLATE_SHEET_NAME}」シートを含むファイルを採用: {os.path.basename(path)}")
            with open(path, "rb") as f:
                return f.read(), "local_file", log
        else:
            log.append(f"「{os.path.basename(path)}」に「{TEMPLATE_SHEET_NAME}」シートが無いためスキップ"
                       f"（シート一覧: {sheetnames}）")

    if not xlsx_paths:
        log.append(
            "同じフォルダに.xlsxが1つも見つかりません。GitHubリポジトリ上でapp.pyと"
            "同じディレクトリにテンプレートxlsxをコミットし、Streamlit Cloud側で"
            "再デプロイ（Reboot）されているか確認してください。"
        )

    return None, "none", log


# ============================================================
# APIキー取得
# ============================================================

def get_api_keys():
    try:
        gemini_key = st.secrets["api_keys"]["gemini"]
        navitime_key = st.secrets["api_keys"]["navitime"]
        return gemini_key, navitime_key
    except KeyError:
        st.error(
            "APIキーが設定されていません。\n\n"
            "Streamlit Cloud の Settings > Secrets に以下の形式で登録してください:\n\n"
            "```toml\n[api_keys]\ngemini = \"YOUR_KEY\"\nnavitime = \"YOUR_KEY\"\n```"
        )
        st.stop()
        return "", ""


# ============================================================
# Gemini リトライラッパー
# ============================================================

def call_gemini_with_retry(client, model, contents, config, retry_status_placeholder=None):
    last_error = None
    for attempt in range(MAX_RETRIES):
        try:
            response = client.models.generate_content(
                model=model, contents=contents, config=config,
            )
            return response.text
        except Exception as e:
            last_error = e
            error_str = str(e)
            is_retryable = any(str(code) in error_str for code in RETRYABLE_STATUS_CODES)
            if not is_retryable:
                raise e
            if attempt >= MAX_RETRIES - 1:
                break
            wait_sec = RETRY_WAIT_SECONDS[attempt]
            if retry_status_placeholder:
                retry_status_placeholder.warning(
                    f"サーバー混雑中... {wait_sec}秒後にリトライ（{attempt+1}/{MAX_RETRIES}）"
                )
            time.sleep(wait_sec)
            if retry_status_placeholder:
                retry_status_placeholder.empty()
    raise last_error


# ============================================================
# ユーティリティ
# ============================================================

def round_up_1000(amount):
    if amount <= 0:
        return 0
    return int(math.ceil(amount / 1000.0) * 1000)


def haversine_km(lat1, lon1, lat2, lon2):
    r = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon / 2) ** 2)
    return 2 * r * math.asin(math.sqrt(a))


def guess_airport_from_address(address):
    mapping = [
        (["沖永良部", "知名町", "和泊町"], "沖永良部空港"),
        (["奄美", "龍郷町"], "奄美空港"),
        (["徳之島", "伊仙町", "天城町"], "徳之島空港"),
        (["与論"], "与論空港"),
        (["石垣"], "新石垣空港"),
        (["宮古"], "宮古空港"),
        (["久米島"], "久米島空港"),
        (["屋久島"], "屋久島空港"),
        (["種子島"], "種子島空港"),
        (["喜界"], "喜界空港"),
        (["沖縄", "那覇", "宜野湾", "浦添"], "那覇空港"),
    ]
    for keywords, airport in mapping:
        if any(kw in address for kw in keywords):
            return airport
    return "最寄り空港"


def parse_json_from_text(text):
    start_idx = text.find('{')
    if start_idx == -1:
        return {}
    depth = 0
    end_idx = start_idx
    for i in range(start_idx, len(text)):
        if text[i] == '{':
            depth += 1
        elif text[i] == '}':
            depth -= 1
            if depth == 0:
                end_idx = i
                break
    if depth != 0:
        match = re.search(r'\{[^{}]*\}', text)
        if match:
            try:
                return json.loads(match.group(0))
            except json.JSONDecodeError:
                return {}
        return {}
    json_str = text[start_idx:end_idx + 1]
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        cleaned = re.sub(r'//.*?\n', '\n', json_str)
        cleaned = re.sub(r',\s*}', '}', cleaned)
        cleaned = re.sub(r',\s*]', ']', cleaned)
        try:
            return json.loads(cleaned)
        except json.JSONDecodeError:
            return {}


# ============================================================
# Excel出力（テンプレート方式）
# ============================================================

def create_excel_report(pattern_data, address, headcount, work_days, origin_station, template_bytes):
    """テンプレートExcelを読み込み、値のみ書き込んで返す"""
    if not template_bytes:
        raise FileNotFoundError(
            "テンプレートExcelが見つかりません。サイドバーの「テンプレートExcelをアップロード」から"
            "テンプレートファイル（交通費シートを含むxlsx）を一度アップロードしてください。"
        )
    try:
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    except Exception as e:
        raise ValueError(f"テンプレートExcelの読み込みに失敗しました（壊れたファイルの可能性）: {e}")

    if TEMPLATE_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"テンプレートExcelに「{TEMPLATE_SHEET_NAME}」シートが見つかりません。"
            f"シート一覧: {wb.sheetnames}"
        )
    ws = wb[TEMPLATE_SHEET_NAME]

    ws['C3'] = address
    ws['C4'] = headcount
    ws['C5'] = work_days

    excel_data = pattern_data.get("excel_data", {})
    time_hours = round(pattern_data.get("time_min", 0) / 60.0, 1)
    stay_note = pattern_data.get("note", "")

    nights = 0
    match_nights = re.search(r'(\d+)泊', stay_note)
    if match_nights:
        nights = int(match_nights.group(1))

    train = excel_data.get("train", {})
    if train.get("used"):
        ws['B9'] = "✓"
        ws['H9'] = train.get("from", "")
        ws['J9'] = train.get("to", "")
        ws['N9'] = train.get("fare", 0)
        ws['O9'] = 1
        ws['R9'] = train.get("time_h", time_hours)
    else:
        ws['B9'] = "-"
        ws['H9'] = ""
        ws['J9'] = ""
        ws['N9'] = 0
        ws['O9'] = 1
        ws['R9'] = ""

    flight = excel_data.get("flight", {})
    if flight.get("used"):
        ws['B10'] = "✓"
        ws['H10'] = flight.get("from", "")
        ws['J10'] = flight.get("to", "")
        ws['N10'] = flight.get("fare", 0)
        ws['O10'] = 1
        ws['R10'] = flight.get("time_h", time_hours)
    else:
        ws['B10'] = "-"
        ws['H10'] = ""
        ws['J10'] = ""
        ws['N10'] = 0
        ws['O10'] = 1
        ws['R10'] = ""

    access = excel_data.get("access", {})
    if access.get("used"):
        ws['B11'] = "✓"
        ws['H11'] = access.get("from", "")
        ws['J11'] = access.get("to", "")
        ws['N11'] = access.get("fare", 0)
        ws['O11'] = 1
        ws['R11'] = ""
    else:
        ws['B11'] = "-"
        ws['H11'] = ""
        ws['J11'] = ""
        ws['N11'] = 0
        ws['O11'] = 1
        ws['R11'] = ""

    rental = excel_data.get("rental", {})
    if rental.get("used"):
        ws['B12'] = "✓"
        ws['N12'] = 0
        ws['O12'] = rental.get("days", 1)
        ws['R12'] = 1
    else:
        ws['B12'] = "-"
        ws['N12'] = 0
        ws['O12'] = 1
        ws['R12'] = ""

    taxi = excel_data.get("taxi", {})
    if taxi.get("used"):
        ws['B13'] = "✓"
        ws['N13'] = taxi.get("fare_per_trip", 0)
    else:
        ws['B13'] = "-"
        ws['N13'] = 0

    hotel = excel_data.get("hotel", {})
    if hotel.get("used"):
        ws['B14'] = "✓"
        ws['O14'] = hotel.get("nights", nights)
    else:
        ws['B14'] = "-"
        ws['O14'] = 0

    ws['B25'] = ws['B9'].value or "-"
    ws['B26'] = ws['B10'].value or "-"
    ws['B27'] = ws['B11'].value or "-"
    ws['B28'] = ws['B12'].value or "-"
    ws['B29'] = ws['B13'].value or "-"
    ws['B30'] = ws['B14'].value or "-"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ============================================================
# STEP 1: Gemini 目的地分析
# ============================================================

def geocode_fallback(address):
    try:
        res = requests.get(GSI_GEOCODE_URL, params={"q": address}, timeout=5)
        if res.status_code == 200 and res.json():
            lon, lat = res.json()[0]["geometry"]["coordinates"]
            return lat, lon
    except Exception:
        pass
    return None, None


def analyze_destination_with_gemini(raw_address, gemini_key, origin_name):
    client = genai.Client(api_key=gemini_key.strip())
    prompt = (
        "以下の住所・施設名について、基本情報をJSON形式で回答してください。\n"
        "Google検索は不要です。あなたの知識のみで回答してください。\n\n"
        f"【対象】: {raw_address}\n\n"
        "以下の形式のJSONテキストのみを回答してください。\n"
        '{\n'
        '  "normalized_address": "正式な住所（都道府県から）",\n'
        '  "dest_lat": 緯度,\n'
        '  "dest_lon": 経度,\n'
        '  "is_island_or_remote": 離島ならtrue/本土はfalse,\n'
        '  "nearest_airport_name": "離島の場合のみ最寄り空港名"\n'
        '}'
    )
    islands = ["沖縄", "奄美", "沖永良部", "石垣", "宮古", "屋久島", "種子島", "徳之島", "与論", "久米島", "喜界"]
    is_island_guess = any(island in raw_address for island in islands)

    fallback_data = {
        "normalized_address": raw_address,
        "is_island_or_remote": is_island_guess,
        "nearest_airport_name": guess_airport_from_address(raw_address) if is_island_guess else "",
        "dest_lat": None, "dest_lon": None,
    }

    retry_placeholder = st.empty()
    try:
        config = types.GenerateContentConfig(temperature=0.1)
        text = call_gemini_with_retry(client=client, model=GEMINI_MODEL, contents=prompt,
                                      config=config, retry_status_placeholder=retry_placeholder)
        if text:
            parsed = parse_json_from_text(text)
            if parsed:
                fallback_data.update(parsed)
    except Exception as e:
        st.warning(f"Gemini API失敗（フォールバック使用）: {e}")
    finally:
        retry_placeholder.empty()

    if not fallback_data.get("dest_lat") or not fallback_data.get("dest_lon"):
        lat, lon = geocode_fallback(fallback_data["normalized_address"])
        fallback_data["dest_lat"] = lat
        fallback_data["dest_lon"] = lon

    if fallback_data.get("is_island_or_remote"):
        guessed = guess_airport_from_address(fallback_data["normalized_address"])
        if not fallback_data.get("nearest_airport_name") or fallback_data["nearest_airport_name"] == "最寄り空港":
            fallback_data["nearest_airport_name"] = guessed

    return fallback_data


def search_flight_fare_with_gemini(raw_address, airport_name, gemini_key, origin_name):
    client = genai.Client(api_key=gemini_key.strip())
    origin_city = "大阪" if "淀屋橋" in origin_name else "東京"
    prompt = (
        "出張旅費算出のためGoogle検索で調査してください。\n\n"
        f"【出発地】: {origin_city}\n"
        f"【到着空港】: {airport_name}\n"
        f"【最終目的地】: {raw_address}\n\n"
        "以下の形式のJSONテキストのみを回答してください。\n"
        '{\n'
        '  "airport_lat": 到着空港の緯度,\n'
        '  "airport_lon": 到着空港の経度,\n'
        f'  "flight_fare_estimate": {origin_city}から{airport_name}への片道普通運賃(円),\n'
        f'  "flight_time_min": {origin_city}から{airport_name}までの片道所要時間(分)\n'
        '}'
    )
    fallback_data = {"airport_lat": None, "airport_lon": None,
                     "flight_fare_estimate": 60000, "flight_time_min": 180}
    retry_placeholder = st.empty()
    try:
        tool = types.Tool(google_search=types.GoogleSearch())
        config = types.GenerateContentConfig(tools=[tool], temperature=0.1)
        text = call_gemini_with_retry(client=client, model=GEMINI_MODEL, contents=prompt,
                                      config=config, retry_status_placeholder=retry_placeholder)
        if text:
            parsed = parse_json_from_text(text)
            if parsed:
                fallback_data.update(parsed)
    except Exception as e:
        st.warning(f"運賃検索失敗（フォールバック使用）: {e}")
    finally:
        retry_placeholder.empty()
    return fallback_data


def get_navitime_fastest_route(start_lat, start_lon, goal_lat, goal_lon, navitime_key,
                               no_flight=False, no_shinkansen=False):
    headers = {"X-RapidAPI-Key": navitime_key.strip(), "X-RapidAPI-Host": NAVITIME_HOST}
    future_day = datetime.now() + timedelta(days=21)
    params = {
        "start": f"{start_lat},{start_lon}",
        "goal": f"{goal_lat},{goal_lon}",
        "start_time": future_day.strftime("%Y-%m-%dT09:00:00"),
        "format": "json"
    }
    unuse_list = []
    if no_flight:
        unuse_list.append("domestic_flight")
    if no_shinkansen:
        unuse_list.append("superexpress_train")
    if unuse_list:
        params["unuse"] = ",".join(unuse_list)

    try:
        res = requests.get(NAVITIME_URL, headers=headers, params=params, timeout=15)
        if res.status_code != 200:
            st.warning(f"NAVITIME エラー: HTTP {res.status_code}")
            return None
        data = res.json()
        items = data.get("items", [])
        if not items:
            st.warning("NAVITIME: ルートが見つかりません。")
            return None

        fastest = min(items, key=lambda x: x.get("summary", {}).get("move", {}).get("time", 99999))
        time_min = fastest.get("summary", {}).get("move", {}).get("time", 0)
        move_info = fastest.get("summary", {}).get("move", {})
        move_types = move_info.get("move_type", [])
        has_superexpress = "superexpress_train" in move_types

        sections = fastest.get("sections", [])

        has_flight = False
        flight_section_idx = -1

        for i, sec in enumerate(sections):
            if sec.get("type") == "move" and sec.get("move", "").lower() in FLIGHT_MOVE_TYPES:
                has_flight = True
                flight_section_idx = i
                break

        pre_flight_fare = 0
        flight_fare = 0
        post_flight_fare = 0
        pre_flight_has_superexpress = False
        post_flight_has_superexpress = False

        start_station_name = None
        end_station_name = None
        end_station_lat = None
        end_station_lon = None
        flight_start_name = None
        flight_end_name = None
        post_flight_start_name = None
        post_flight_end_name = None

        if has_flight:
            for i, sec in enumerate(sections):
                sec_type = sec.get("type", "")
                m_type = sec.get("move", "").lower()

                if sec_type == "point":
                    if not start_station_name:
                        start_station_name = sec.get("name")
                    if i < flight_section_idx and sections[i + 1] if i + 1 < len(sections) else None:
                        pass
                    if "goal" not in sec.get("name", "").lower():
                        end_station_name = sec.get("name")
                        if "coord" in sec:
                            end_station_lat = sec["coord"].get("lat")
                            end_station_lon = sec["coord"].get("lon")

                if sec_type == "move":
                    transport = sec.get("transport", {})
                    sec_fare = 0
                    # [修正] unit_3 ではなく unit_1（運賃(unit_0)と組み合わせて使う
                    # 特急・新幹線料金）を使う。unit_3 は別の料金体系(切符種別)の値で、
                    # unit_0 と単純に足すと無関係な組み合わせになり過大になる。
                    if transport and "fare" in transport and isinstance(transport["fare"], dict):
                        sec_fare = int(transport["fare"].get("unit_0", 0))
                        sec_unit_1 = int(transport["fare"].get("unit_1", 0))
                    else:
                        sec_unit_1 = 0

                    if i < flight_section_idx:
                        pre_flight_fare += sec_fare + sec_unit_1
                        if m_type in ("superexpress_train",):
                            pre_flight_has_superexpress = True
                    elif i == flight_section_idx:
                        flight_fare = sec_fare
                    else:
                        post_flight_fare += sec_fare + sec_unit_1
                        if m_type in ("superexpress_train",):
                            post_flight_has_superexpress = True

            for i in range(flight_section_idx - 1, -1, -1):
                if sections[i].get("type") == "point":
                    flight_start_name = sections[i].get("name")
                    break
            for i in range(flight_section_idx + 1, len(sections)):
                if sections[i].get("type") == "point":
                    flight_end_name = sections[i].get("name")
                    break

            found_flight_end = False
            for i in range(flight_section_idx + 1, len(sections)):
                if sections[i].get("type") == "point":
                    if not found_flight_end:
                        found_flight_end = True
                        post_flight_start_name = sections[i].get("name")
                    else:
                        if "goal" not in sections[i].get("name", "").lower():
                            post_flight_end_name = sections[i].get("name")
                            if "coord" in sections[i]:
                                end_station_lat = sections[i]["coord"].get("lat")
                                end_station_lon = sections[i]["coord"].get("lon")

        else:
            fare_dict = move_info.get("fare", {})
            reference_fare = fare_dict.get("reference_fare", {}) if isinstance(fare_dict, dict) else {}

            # [修正] NAVITIME自身が算出した合計運賃(reference_fare)を最優先で使う。
            # これが最も信頼できる「実際に乗車した場合の合計運賃」。
            lowest_ic = reference_fare.get("lowest_total_ic") if isinstance(reference_fare, dict) else None
            lowest_ticket = reference_fare.get("lowest_total_ticket") if isinstance(reference_fare, dict) else None

            if lowest_ic or lowest_ticket:
                pre_flight_fare = int(lowest_ic or lowest_ticket)
            else:
                # フォールバック: 運賃(unit_0) + 特急・新幹線料金(unit_1) の組み合わせ。
                # 旧コードは unit_0 + unit_3 を合算しており、無関係な料金体系同士を
                # 足してしまい実際より高い金額になるバグがあった。
                fare_unit_0 = fare_dict.get("unit_0", 0) if isinstance(fare_dict, dict) else 0
                fare_unit_1 = fare_dict.get("unit_1", 0) if isinstance(fare_dict, dict) else 0
                pre_flight_fare = int(fare_unit_0 + fare_unit_1) if has_superexpress else int(fare_unit_0)

            for sec in sections:
                if sec.get("type") == "point":
                    if not start_station_name:
                        start_station_name = sec.get("name")
                    if "goal" not in sec.get("name", "").lower():
                        end_station_name = sec.get("name")
                        if "coord" in sec:
                            end_station_lat = sec["coord"].get("lat")
                            end_station_lon = sec["coord"].get("lon")

        last_walk_min = 0
        if sections and sections[-1].get("type") == "move":
            last_walk_min = sections[-1].get("time", 0)

        result = {
            "has_flight": has_flight,
            "time_min": time_min,
            "pre_flight_fare": int(pre_flight_fare),
            "flight_fare": int(flight_fare),
            "post_flight_fare": int(post_flight_fare),
            "total_fare": int(pre_flight_fare + flight_fare + post_flight_fare),
            "last_walk_min": last_walk_min,
            "start_station": start_station_name or "出発駅",
            "end_station": end_station_name or "到着駅",
            "end_station_lat": end_station_lat,
            "end_station_lon": end_station_lon,
            "flight_start": flight_start_name,
            "flight_end": flight_end_name,
            "post_flight_start": post_flight_start_name,
            "post_flight_end": post_flight_end_name or end_station_name,
        }

        if DEBUG_MODE:
            with st.expander("🔍 [DEBUG] NAVITIME詳細", expanded=False):
                st.json(params)
                st.json(move_info)
                st.markdown(f"**move_type:** `{move_types}` | **superexpress:** `{has_superexpress}`")
                if has_flight:
                    st.markdown(f"**前区間(電車):** {pre_flight_fare:,} 円")
                    st.markdown(f"**飛行機:** {flight_fare:,} 円")
                    st.markdown(f"**後区間(電車/新幹線):** {post_flight_fare:,} 円")
                    st.markdown(f"**空港:** {flight_start_name} → {flight_end_name}")
                    st.markdown(f"**後区間駅:** {post_flight_start_name} → {post_flight_end_name}")
                else:
                    st.markdown(f"**運賃合計(片道):** {pre_flight_fare:,} 円")
                st.json(result)
        return result
    except Exception as e:
        st.warning(f"NAVITIME エラー: {e}")
        return None


def build_best_route_patterns(current_st_name, ai_info, navitime_route,
                              headcount, work_days, no_rental=False, no_taxi=False):
    patterns = []

    if navitime_route:
        has_flight = navitime_route["has_flight"]
        time_min = navitime_route["time_min"]
        dest_lat, dest_lon = ai_info.get("dest_lat"), ai_info.get("dest_lon")
        e_lat, e_lon = navitime_route.get("end_station_lat"), navitime_route.get("end_station_lon")

        if e_lat and e_lon and dest_lat and dest_lon:
            base_taxi_km = haversine_km(e_lat, e_lon, dest_lat, dest_lon) * 1.3
            base_taxi_km = 0.0 if base_taxi_km <= 1.2 else max(base_taxi_km, 1.5)
        else:
            base_taxi_km = (0.0 if navitime_route["last_walk_min"] <= TAXI_WALK_THRESHOLD_MIN
                            else max(navitime_route["last_walk_min"] * 0.08, 1.5))

        if has_flight:
            selected_mode = "flight"
            pre_flight_fare = navitime_route["pre_flight_fare"]
            flight_fare = navitime_route["flight_fare"]
            post_flight_fare = navitime_route["post_flight_fare"]
            origin_airport = navitime_route["flight_start"] or "出発空港"
            dest_airport = navitime_route["flight_end"] or "到着空港"
            end_st = navitime_route["end_station"]
            post_start = navitime_route.get("post_flight_start") or dest_airport
            post_end = navitime_route.get("post_flight_end") or end_st

            route_title = f"✈️ 飛行機ルート ({dest_airport}経由)"
            display_route_str = f"{current_st_name} ➔ {origin_airport} ➔ {dest_airport} ➔ {end_st} ➔ 目的地"

            # 行9=「電車・新幹線（往復）」＝出発駅→出発空港（pre_flight）
            # 行10=「飛行機（往復）」
            # 行11=（到着側アクセス電車）＝到着空港→最終駅（post_flight）
            # という並び順がテンプレートの実際のレイアウトなので、
            # train(行9)にはpre_flight、access(行11)にはpost_flightを入れる
            #
            # [修正] さらに、NAVITIMEの片道検索結果である
            # pre_flight_fare / flight_fare / post_flight_fare を
            # そのままExcelのN列に書き込むと「片道」の金額になってしまう。
            # テンプレートの行ラベルはいずれも「（往復）」であり、
            # Q列の合計式(=M×O×P)は「Mに往復1人分の運賃が入っている」前提で
            # 人数(P)しか掛けていないため、往復分の×2はここで明示的に行う。
            excel_data = {
                "train": {
                    "used": pre_flight_fare > 0,
                    "from": current_st_name.replace("駅", ""),
                    "to": origin_airport,
                    "fare": pre_flight_fare * 2,
                    "time_h": "",
                },
                "flight": {
                    "used": True,
                    "from": origin_airport,
                    "to": dest_airport,
                    "fare": flight_fare * 2,
                    "time_h": time_hours if (time_hours := round(time_min / 60.0, 1)) else "",
                },
                "access": {
                    "used": post_flight_fare > 0,
                    "from": post_start,
                    "to": post_end,
                    "fare": post_flight_fare * 2,
                },
            }
            is_ai_fare = False
        else:
            selected_mode = "train"
            end_st = navitime_route["end_station"]
            total_fare = navitime_route["total_fare"]
            route_title = f"🚄 電車ルート ({end_st}着)"
            display_route_str = f"{current_st_name} ➔ {end_st} ➔ 目的地"

            # [修正] total_fare はNAVITIMEの片道検索結果（片道運賃）。
            # 行9「電車・新幹線（往復）」に書き込むため、明示的に2倍して往復運賃にする。
            excel_data = {
                "train": {
                    "used": True,
                    "from": current_st_name,
                    "to": end_st,
                    "fare": total_fare * 2,
                    "time_h": round(time_min / 60.0, 1),
                },
                "flight": {"used": False},
                "access": {"used": False},
            }
            is_ai_fare = False
    else:
        selected_mode = "flight"
        airport_name = ai_info.get("nearest_airport_name", "最寄り空港")
        airport_lat = ai_info.get("airport_lat", 0)
        airport_lon = ai_info.get("airport_lon", 0)
        dest_lat = ai_info.get("dest_lat", 0)
        dest_lon = ai_info.get("dest_lon", 0)
        airport_to_dest_km = 15.0
        if airport_lat and airport_lon and dest_lat and dest_lon:
            airport_to_dest_km = max(haversine_km(airport_lat, airport_lon, dest_lat, dest_lon) * 1.3, 1.0)
        time_min = int(60 + ai_info.get("flight_time_min", 150) + (airport_to_dest_km / 40.0 * 60))
        flight_fare = int(ai_info.get("flight_fare_estimate", 60000) * 1.3)
        origin_airport = "伊丹空港" if "淀屋橋" in current_st_name else "羽田空港"
        access_fare = 1500 if "大宮" in current_st_name else 500
        end_st = airport_name
        base_taxi_km = airport_to_dest_km

        route_title = f"✈️ 飛行機ルート ({airport_name}利用)"
        display_route_str = f"{current_st_name} ➔ {origin_airport} ➔ {airport_name} ➔ 目的地"

        # [修正] flight_fare / access_fare も片道相当の値なので、
        # Excelの「（往復）」欄に書き込む際は2倍する。
        excel_data = {
            "train": {"used": False},
            "flight": {
                "used": True,
                "from": origin_airport,
                "to": airport_name,
                "fare": flight_fare * 2,
                "time_h": round(time_min / 60.0, 1),
            },
            "access": {
                "used": access_fare > 0,
                "from": current_st_name.replace("駅", ""),
                "to": origin_airport,
                "fare": access_fare * 2,
            },
        }
        is_ai_fare = True
        post_flight_fare = 0
        pre_flight_fare = access_fare

    travel_hours = time_min / 60.0
    if selected_mode == "flight" or travel_hours >= 4.0:
        nights = work_days + 1
        stay_note = "前泊/後泊想定（" + str(nights) + "泊）"
    elif travel_hours >= 2.5:
        nights = work_days
        stay_note = "宿泊想定（" + str(nights) + "泊）"
    else:
        nights = max(work_days - 1, 0)
        stay_note = "標準（" + str(nights) + "泊）"

    hotel_cost = round_up_1000(HOTEL_COST_PER_NIGHT_PER_PERSON * headcount * nights)
    rental_days = work_days + (1 if "前泊" in stay_note else 0)
    rental_car_total = round_up_1000(RENTAL_CAR_COST_PER_DAY * rental_days)
    taxi_one_way = base_taxi_km * TAXI_FARE_PER_KM
    taxi_trips = nights + 1
    taxi_total = round_up_1000(taxi_one_way * 2 * taxi_trips)

    base_transport = {}
    if selected_mode == "flight":
        base_transport["航空券費用(往復・人数分・flex)"] = round_up_1000(flight_fare * 2 * headcount)
        if navitime_route and navitime_route.get("pre_flight_fare", 0) > 0:
            base_transport["アクセス電車運賃(往復・人数分)"] = round_up_1000(navitime_route["pre_flight_fare"] * 2 * headcount)
        elif not navitime_route and access_fare > 0:
            base_transport["アクセス電車運賃(往復・人数分)"] = round_up_1000(access_fare * 2 * headcount)
        if navitime_route and navitime_route.get("post_flight_fare", 0) > 0:
            base_transport["到着後電車・新幹線(往復・人数分)"] = round_up_1000(navitime_route["post_flight_fare"] * 2 * headcount)
    else:
        base_transport["電車・新幹線運賃(往復・人数分)"] = round_up_1000(navitime_route["total_fare"] * 2 * headcount)

    b_taxi = dict(base_transport)
    if taxi_total > 0:
        b_taxi["現地タクシー運賃(往復x" + str(taxi_trips) + "回)"] = taxi_total
    b_taxi["宿泊費"] = hotel_cost
    taxi_sum = sum(b_taxi.values())

    b_rental = dict(base_transport)
    b_rental["レンタカー費用(12000円x" + str(rental_days) + "日)"] = rental_car_total
    b_rental["宿泊費"] = hotel_cost
    rental_sum = sum(b_rental.values())

    if no_taxi and no_rental:
        final_breakdown = dict(base_transport)
        final_breakdown["宿泊費"] = hotel_cost
        final_cost = sum(final_breakdown.values())
        final_type = "walk"
        final_name = route_title + " (徒歩)"
        recommend_msg = "最安ルート (徒歩前提)"
        excel_data["rental"] = {"used": False}
        excel_data["taxi"] = {"used": False}
    elif no_rental:
        final_breakdown = b_taxi
        final_cost = taxi_sum
        final_type = "taxi"
        final_name = route_title + (" (徒歩)" if taxi_total == 0 else " + タクシー")
        recommend_msg = "最安ルート (タクシー)"
        excel_data["rental"] = {"used": False}
        excel_data["taxi"] = {"used": taxi_total > 0, "fare_per_trip": round_up_1000(taxi_one_way * 2)}
    elif no_taxi:
        final_breakdown = b_rental
        final_cost = rental_sum
        final_type = "rental"
        final_name = route_title + " + レンタカー"
        recommend_msg = "最安ルート (レンタカー)"
        excel_data["rental"] = {"used": True, "days": rental_days}
        excel_data["taxi"] = {"used": False}
    else:
        if rental_sum < taxi_sum:
            final_breakdown = b_rental
            final_cost = rental_sum
            final_type = "rental"
            final_name = route_title + " + レンタカー"
            recommend_msg = "最安 (レンタカー・タクシーより " + f"{taxi_sum - rental_sum:,}" + " 円お得)"
            excel_data["rental"] = {"used": True, "days": rental_days}
            excel_data["taxi"] = {"used": False}
        else:
            final_breakdown = b_taxi
            final_cost = taxi_sum
            final_type = "taxi"
            final_name = route_title + (" (徒歩)" if taxi_total == 0 else " + タクシー")
            recommend_msg = "最安ルート (タクシー)"
            excel_data["rental"] = {"used": False}
            excel_data["taxi"] = {"used": taxi_total > 0, "fare_per_trip": round_up_1000(taxi_one_way * 2)}

    excel_data["hotel"] = {"used": nights > 0, "nights": nights}

    if is_ai_fare:
        recommend_msg += " / AI相場(1.3倍マージン)"

    patterns.append({
        "type": final_type, "name": final_name, "time_min": time_min,
        "cost": final_cost, "breakdown": final_breakdown, "note": stay_note,
        "routes": {}, "display_route": display_route_str,
        "recommend_reason": recommend_msg, "is_recommended": True,
        "excel_data": excel_data,
    })
    return patterns


# ============================================================
# Streamlit UI
# ============================================================

st.set_page_config(page_title="交通費・出張見積もりアプリ", page_icon="🚗", layout="wide")
st.title("🚗 交通費・出張見積もりアプリ")

if DEBUG_MODE:
    st.caption("🐛 デバッグモード ON")

gemini_api_key, navitime_api_key = get_api_keys()

with st.sidebar:
    st.markdown("### 📄 Excelテンプレート")
    tpl_file = st.file_uploader(
        "テンプレートExcelを手動アップロード（未設定の場合や上書きしたい場合）",
        type=["xlsx"],
        key="template_uploader",
    )
    if tpl_file is not None:
        st.session_state["template_bytes"] = tpl_file.getvalue()

    if st.session_state.get("template_bytes"):
        st.caption("✅ 手動アップロードしたテンプレートを使用中")

    _resolved_bytes, _source, _log = resolve_template_bytes()
    if _source == "github":
        st.caption("🔗 GitHubから取得済み")
    elif _source == "local_file":
        st.caption("💻 ローカルファイルから取得済み")
    elif _source == "manual_upload":
        pass  # 上ですでに表示済み
    else:
        st.caption("⚠️ テンプレートが見つかりません")

    with st.expander("🔍 テンプレート取得ログ", expanded=(_source == "none")):
        for line in _log:
            st.caption(f"・{line}")

col1, col2 = st.columns([2, 1])
with col1:
    address_input = st.text_input("目的地（住所や施設名）", "")
with col2:
    station_choice = st.selectbox("出発拠点", ["淀屋橋駅", "大宮駅", "両方比較"], index=0)

col_a, col_b = st.columns(2)
with col_a:
    headcount = st.number_input("作業人数（人）", min_value=1, value=2)
with col_b:
    work_days = st.number_input("現地作業日数（日）", min_value=1, value=2)

st.markdown("**使用しない交通手段:**")
col_c1, col_c2, col_c3, col_c4 = st.columns(4)
with col_c1:
    no_flight = st.checkbox("飛行機を使用しない", value=False)
with col_c2:
    no_shinkansen = st.checkbox("新幹線を使用しない", value=False)
with col_c3:
    no_rental = st.checkbox("レンタカーを使用しない", value=False)
with col_c4:
    no_taxi = st.checkbox("タクシーを使用しない", value=False)

st.markdown("---")

if st.button("見積もりを計算する", type="primary"):
    if not address_input.strip():
        st.warning("目的地を入力してください。")
        st.stop()

    if no_flight and no_shinkansen:
        st.info("飛行機・新幹線の両方を除外。在来線のみで検索します。")

    stations = (STATION_COORDS if station_choice == "両方比較"
                else {station_choice: STATION_COORDS[station_choice]})

    for current_st_name, (current_lat, current_lon) in stations.items():
        st.markdown(f"### 出発地: 【{current_st_name}】")

        with st.spinner(f"AIが {address_input} を分析中..."):
            ai_info = analyze_destination_with_gemini(address_input, gemini_api_key, current_st_name)

        if not ai_info.get("dest_lat") or not ai_info.get("dest_lon"):
            st.error("座標取得失敗。住所を詳しく入力してください。")
            continue

        st.success(f"検索地点: {ai_info.get('normalized_address', address_input)}")

        if DEBUG_MODE:
            with st.expander("🔍 [DEBUG] Gemini結果", expanded=False):
                st.json(ai_info)

        train_route = None
        if ai_info.get("is_island_or_remote", False) and not no_flight:
            st.info("離島判定 → 飛行機ルートを適用")
            airport_name = ai_info.get("nearest_airport_name", "最寄り空港")
            with st.spinner(f"{airport_name} への運賃を検索中..."):
                fare_info = search_flight_fare_with_gemini(
                    address_input, airport_name, gemini_api_key, current_st_name)
            ai_info.update(fare_info)
            if not ai_info.get("airport_lat"):
                ai_info["airport_lat"] = ai_info.get("dest_lat", 24.3964)
                ai_info["airport_lon"] = ai_info.get("dest_lon", 124.2450)
        elif ai_info.get("is_island_or_remote", False) and no_flight:
            st.warning("離島ですが飛行機除外。")
            with st.spinner("NAVITIMEで検索中..."):
                train_route = get_navitime_fastest_route(
                    current_lat, current_lon, ai_info["dest_lat"], ai_info["dest_lon"],
                    navitime_api_key, no_flight=no_flight, no_shinkansen=no_shinkansen)
        else:
            with st.spinner("NAVITIMEで最速ルートを検索中..."):
                train_route = get_navitime_fastest_route(
                    current_lat, current_lon, ai_info["dest_lat"], ai_info["dest_lon"],
                    navitime_api_key, no_flight=no_flight, no_shinkansen=no_shinkansen)

        patterns = build_best_route_patterns(
            current_st_name, ai_info, train_route, headcount, work_days,
            no_rental=no_rental, no_taxi=no_taxi)

        for p in patterns:
            with st.expander(f"{p['name']} — 合計 {p['cost']:,} 円", expanded=p["is_recommended"]):
                st.success(p['recommend_reason'])
                st.write(f"**ルート:** {p['display_route']}")
                st.write(f"**片道:** 約 {p['time_min']} 分")
                st.write(f"**宿泊:** {p['note']}")
                st.write("**内訳:**")
                for item, amt in p["breakdown"].items():
                    st.write(f"　・{item}: **{amt:,}** 円")

                if DEBUG_MODE:
                    with st.expander("🔍 [DEBUG] excel_data", expanded=False):
                        st.json(p.get("excel_data", {}))

                _tpl_bytes, _tpl_source, _tpl_log = resolve_template_bytes()
                try:
                    excel_bytes = create_excel_report(
                        p, ai_info.get('normalized_address', address_input),
                        headcount, work_days, current_st_name,
                        template_bytes=_tpl_bytes)
                    st.download_button(
                        label="Excel見積書をダウンロード",
                        data=excel_bytes,
                        file_name=f"交通費見積_{current_st_name}発_{address_input[:10]}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{current_st_name}_{p['type']}")
                except (FileNotFoundError, ValueError) as e:
                    st.error(f"Excel出力エラー: {e}")
                    with st.expander("🔍 テンプレート取得ログ（原因確認用）", expanded=True):
                        for line in _tpl_log:
                            st.caption(f"・{line}")
        st.markdown("---")
